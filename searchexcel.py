from tkinter import *
import tkinter as tk
import openpyxl,xlrd
from openpyxl import Workbook


main=Tk()
main.title("Login Form")
main.geometry("800x500")
main.config(highlightbackground="black")

excel_path =r".\Backened_Data.xlsx"
def submit():
    search_empid=empid.get()

    name.configure(state=tk.NORMAL)
    user.configure(state=tk.NORMAL)
    emailentry.configure(state=tk.NORMAL)
    gender.configure(state=tk.NORMAL)
    subscription.configure(state=tk.NORMAL)

    name.delete(0, 'end')
    user.delete(0, 'end')
    emailentry.delete(0, 'end')
    gender.delete(0, 'end')
    subscription.delete(0, 'end')

    file =openpyxl.load_workbook(excel_path)
    sheet=file['Sheet']

    for cell in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=1,max_col=7,values_only=True):

        if cell[0] ==int(search_empid):
            name.insert(0, cell[1])
            user.insert(0, cell[2])
            emailentry.insert(0, cell[4])
            gender.insert(0, cell[5])
            subscription.insert(0, cell[6])

            name.configure(state=tk.DISABLED)
            user.configure(state=tk.DISABLED)
            emailentry.configure(state=tk.DISABLED)
            gender.configure(state=tk.DISABLED)
            subscription.configure(state=tk.DISABLED)

frame1 = LabelFrame(main, text = 'Login Details:').pack(expand = 'yes', fill = 'both')

Label(frame1,text="Emp ID:").place(x=50,y=60)
Label(frame1,text="Name:").place(x=50,y=100)
Label(frame1,text="Username:").place(x=50,y=130)

Label(main,text="Mail ID:").place(x=50,y=160)

frame2 = LabelFrame(main, text = 'Other Details:').pack(expand = 'yes', fill = 'both')

empid =Entry(frame1)
empid.place(x=250,y=60)
name = Entry(frame1)
name.place(x=250,y=100)
user =Entry(frame1)
user.place(x=250,y=130)
emailentry =Entry(frame1)
emailentry.place(x=250,y=130,width=250)


label_3 = Label(frame2, text="Gender",width=20,font=("bold", 10))
label_3.place(x=50,y=300)
gender =Entry(frame2)
gender.place(x=200,y=300)


label_4 = Label(frame2, text="Subscription",width=20,font=("bold", 10))
label_4.place(x=50,y=360)
subscription =Entry(frame2)
subscription.place(x=200,y=360)

name.configure(state=tk.DISABLED)
user.configure(state=tk.DISABLED)
emailentry.configure(state=tk.DISABLED)
gender.configure(state=tk.DISABLED)
subscription.configure(state=tk.DISABLED)

Button(frame2,text = "Search",command= submit).place(x=400,y=400)

main.mainloop()